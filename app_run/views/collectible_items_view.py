from openpyxl.reader.excel import load_workbook
from rest_framework import status
from rest_framework.generics import ListAPIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView

from ..models import CollectibleItem
from ..serializers.serializers import CollectibleItemValidator, FileUploadSerializer, CollectibleItemSerializer


class CollectibleItemView(ListAPIView):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer


class UploadFileView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        input_serializer = FileUploadSerializer(data=request.data)
        input_serializer.is_valid(raise_exception=True)
        file_object = input_serializer.validated_data["file"]

        workbook = load_workbook(file_object)
        worksheet = workbook.active

        headers = CollectibleItemValidator.Meta.fields

        wrong_rows = list()

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_serializer = CollectibleItemValidator(data=dict(zip(headers, row)))
            if row_serializer.is_valid(raise_exception=False):
                row_serializer.save()
            else:
                wrong_rows.append(list(row))

        return Response(wrong_rows, status=status.HTTP_200_OK)
